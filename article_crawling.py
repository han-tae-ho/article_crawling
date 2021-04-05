import time
import requests
from bs4 import BeautifulSoup
import os
import openpyxl as xl

#데이터 쌓을 날짜 입력
DATES = "20210405"

#데이터 쌓을 언론사 입력
press = "경향신문"

#url
url = "https://news.naver.com/main/list.nhn?mode=LPOD&mid=sec"

# 언론사 고유번호
press_codes = {
    "경향신문" : "032",
    "국민일보" : "005",
    "동아일보" : "020",
    "문화일보" : "021",
    "서울신문" : "081",
    "세계일보" : "022",
    "조선일보" : "023",
    "중앙일보" : "025",
    "한겨례" : "028",
    "한국일보" : "469"
}

# 네이버에서 크롤링 차단 http://www.useragentstring.com/ 에서 유저키 발급 후 시도
headers = {'user-agent' : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.114 Safari/537.36"}

# 데이터쌓을 엑셀 지정
wb = xl.load_workbook('crawling_data.xlsx')
sheet = wb['Sheet1']
max_row = sheet.max_row

#### 크롤링
for page in range(1,7):
    params = {
        'oid' : press_codes[press],
        'date' : DATES,
        'page' : page
    }

    html = requests.get(url, headers=headers, params = params).text
    print(requests.get(url, headers=headers, params = params).url)
    soup = BeautifulSoup(html, 'html.parser')


    for href in soup.find("div", class_="newsflash_body").find_all("li"):
        # url 저장
        url_cell = sheet.cell(max_row + 1, 1)
        url_cell.value = href.find("a")["href"]

        # title 저장
        try:
            title_cell = sheet.cell(max_row + 1, 2)
            title_cell.value = href.find("img")['alt']

        except TypeError:
            title_cell = sheet.cell(max_row + 1, 2)
            title_cell.value = href.find("a").text

        max_row += 1

# 엑셀 데이터 저장
wb.save('crawling_data.xlsx')